"""Cross-session experiment monitoring Excel log."""

from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path
from typing import List, Optional

logger = logging.getLogger(__name__)


class MainExperimentMonitor:
    """Creates/updates MAIN_experiment_monitoring.xlsx in the output base directory.

    Appends one row per session with summary information.
    This file persists across sessions and is never overwritten.
    """

    HEADER = [
        "Date",
        "Start Time",
        "End Time",
        "Status",
        "Participant Names",
        "Shapes",
        "Repetitions",
        "Shape Reps Per SubSession",
        "Camera Settings Summary",
        "Output Session Folder",
    ]

    def __init__(self, output_base_dir: str):
        self._path = Path(output_base_dir) / "MAIN_experiment_monitoring.xlsx"

    def log_session(
        self,
        start_time: datetime,
        end_time: datetime,
        status: str,
        participants: List[str],
        shapes: List[str],
        repetitions: int,
        shape_reps_per_subsession: int,
        camera_summary: str,
        session_folder: str,
    ) -> None:
        """Append a session summary row to the monitoring Excel file."""
        try:
            from openpyxl import Workbook, load_workbook

            if self._path.exists():
                wb = load_workbook(str(self._path))
                ws = wb.active
            else:
                self._path.parent.mkdir(parents=True, exist_ok=True)
                wb = Workbook()
                ws = wb.active
                ws.title = "Experiment Monitor"
                ws.append(self.HEADER)
                for col_idx, header in enumerate(self.HEADER, 1):
                    ws.column_dimensions[
                        chr(64 + col_idx) if col_idx <= 26
                        else "A" + chr(64 + col_idx - 26)
                    ].width = max(len(header) + 2, 18)

            ws.append([
                start_time.strftime("%Y-%m-%d"),
                start_time.strftime("%H:%M:%S"),
                end_time.strftime("%H:%M:%S"),
                status,
                ", ".join(participants),
                ", ".join(shapes),
                repetitions,
                shape_reps_per_subsession,
                camera_summary,
                session_folder,
            ])

            wb.save(str(self._path))
            logger.info("Session logged to %s", self._path)

        except ImportError:
            logger.warning("openpyxl not available — skipping monitor log")
        except Exception as e:
            logger.error("Failed to log session to monitor: %s", e)
